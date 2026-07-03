import re
import streamlit as st
import pandas as pd
from datetime import date, datetime, timedelta
from typing import Optional
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Калькулятор пеней",
    page_icon="⚖️",
    layout="wide",
)

# ─── КЛЮЧЕВАЯ СТАВКА ЦБ РФ ───────────────────────────────────────────────────
# Источник: https://cbr.ru/hd_base/KeyRate/
# Актуально на 30.06.2026. При изменении ставки — обновите таблицу на вкладке «Расчёт».
CBR_RATES: list[tuple[date, float]] = [
    # 2013
    (date(2013,  9, 13),  5.50),
    # 2014
    (date(2014,  3,  3),  7.00),
    (date(2014,  4, 28),  7.50),
    (date(2014,  7, 28),  8.00),
    (date(2014, 11,  5),  9.50),
    (date(2014, 12, 12), 10.50),
    (date(2014, 12, 16), 17.00),
    # 2015
    (date(2015,  2,  2), 15.00),
    (date(2015,  3, 16), 14.00),
    (date(2015,  5,  5), 12.50),
    (date(2015,  6, 16), 11.50),
    (date(2015,  8,  3), 11.00),
    # 2016
    (date(2016,  6, 14), 10.50),
    (date(2016,  9, 19), 10.00),
    # 2017
    (date(2017,  3, 27),  9.75),
    (date(2017,  5,  2),  9.25),
    (date(2017,  6, 19),  9.00),
    (date(2017,  9, 18),  8.50),
    (date(2017, 10, 30),  8.25),
    (date(2017, 12, 18),  7.75),
    # 2018
    (date(2018,  2, 12),  7.50),
    (date(2018,  3, 26),  7.25),
    (date(2018,  9, 17),  7.50),
    (date(2018, 12, 17),  7.75),
    # 2019
    (date(2019,  6, 17),  7.50),
    (date(2019,  7, 29),  7.25),
    (date(2019,  9,  9),  7.00),
    (date(2019, 10, 28),  6.50),
    (date(2019, 12, 16),  6.25),
    # 2020
    (date(2020,  2, 10),  6.00),
    (date(2020,  4, 27),  5.50),
    (date(2020,  6, 22),  4.50),
    (date(2020,  7, 27),  4.25),
    # 2021
    (date(2021,  3, 22),  4.50),
    (date(2021,  4, 26),  5.00),
    (date(2021,  6, 15),  5.50),
    (date(2021,  7, 26),  6.50),
    (date(2021,  9, 13),  6.75),
    (date(2021, 10, 25),  7.50),
    (date(2021, 12, 20),  8.50),
    # 2022
    (date(2022,  2, 14),  9.50),
    (date(2022,  2, 28), 20.00),  # экстренное повышение
    (date(2022,  4, 11), 17.00),
    (date(2022,  5,  4), 14.00),
    (date(2022,  5, 27), 11.00),
    (date(2022,  6, 14),  9.50),
    (date(2022,  7, 25),  8.00),
    (date(2022,  9, 19),  7.50),
    # 2023
    (date(2023,  7, 24),  8.50),
    (date(2023,  8, 15), 12.00),
    (date(2023,  9, 18), 13.00),
    (date(2023, 10, 30), 15.00),
    (date(2023, 12, 18), 16.00),
    # 2024
    (date(2024,  7, 29), 18.00),   # решение 26.07.2024, вступило в силу 29.07 (cbr.ru)
    (date(2024,  9, 16), 19.00),   # решение 13.09.2024, вступило в силу 16.09 (cbr.ru)
    (date(2024, 10, 28), 21.00),   # решение 25.10.2024, вступило в силу 28.10 (cbr.ru: 25.10=19%, 28.10=21%)
    # 2025 — ставка 21% удерживалась до 08.06.2025, затем цикл снижения
    (date(2025,  6,  9), 20.00),   # решение 06.06.2025, вступило 09.06 (cbr.ru)
    (date(2025,  7, 28), 18.00),   # решение 25.07.2025, вступило 28.07 (cbr.ru)
    (date(2025,  9, 15), 17.00),   # решение 12.09.2025, вступило 15.09 (cbr.ru)
    (date(2025, 10, 27), 16.50),   # решение 24.10.2025, вступило 27.10 (cbr.ru)
    (date(2025, 12, 22), 16.00),   # решение 19.12.2025, вступило 22.12 (cbr.ru)
    # 2026
    (date(2026,  2, 16), 15.50),   # решение 13.02.2026, вступило 16.02 (cbr.ru)
    (date(2026,  3, 23), 15.00),   # решение 20.03.2026, вступило 23.03 (cbr.ru)
    (date(2026,  4, 27), 14.50),   # решение 24.04.2026, вступило 27.04 (cbr.ru)
    (date(2026,  6, 22), 14.25),   # решение 19.06.2026, вступило 22.06 (cbr.ru)
]


def get_cbr_rate(d: date, rates: list[tuple[date, float]]) -> float:
    applicable = [(dt, r) for dt, r in rates if dt <= d]
    return max(applicable, key=lambda x: x[0])[1] if applicable else 0.0


# ─── РАСЧЁТ ПЕНЕЙ ─────────────────────────────────────────────────────────────
def calculate_penalties(
    schedule: list[tuple[date, float]],
    payments: list[tuple[date, float]],
    cbr_rates: list[tuple[date, float]],
    calc_date: date,
) -> list[dict]:
    if not schedule:
        return []

    # Все узловые даты: плановые платежи, фактические оплаты, изменения ставки ЦБ
    events: set[date] = set()
    for d, _ in schedule:
        events.add(d)
    for d, _ in payments:
        events.add(d)
    for d, _ in cbr_rates:
        if d <= calc_date:
            events.add(d)
    events.add(calc_date)

    start = min(d for d, _ in schedule)
    sorted_events = sorted(e for e in events if e >= start and e <= calc_date)

    plan_by_date: dict[date, float] = {}
    for d, a in schedule:
        plan_by_date[d] = plan_by_date.get(d, 0.0) + a

    pay_by_date: dict[date, float] = {}
    for d, a in payments:
        pay_by_date[d] = pay_by_date.get(d, 0.0) + a

    cum_planned = 0.0
    cum_actual = 0.0
    results: list[dict] = []

    prev_date: Optional[date] = None
    prev_overdue = 0.0
    prev_rate = 0.0

    for curr_date in sorted_events:
        planned_today = plan_by_date.get(curr_date, 0.0)

        if prev_date is not None and prev_overdue > 0.0 and prev_rate > 0.0:
            # Если сегодня наступает плановый платёж — старый долг продолжает
            # начисляться в этот день (пени за новый долг пойдут с завтра).
            # Поэтому закрываем период ВКЛЮЧАЯ curr_date (end = curr_date + 1).
            # Для прочих событий (ставка ЦБ, фактическая оплата) — не включаем.
            period_end = curr_date + timedelta(days=1) if planned_today > 0 else curr_date
            days = (period_end - prev_date).days
            if days > 0:
                penalty = prev_overdue * (1 / 300) * (prev_rate / 100) * days
                results.append({
                    "date_from": prev_date,
                    "date_to":   period_end - timedelta(days=1),
                    "days":      days,
                    "debt":      prev_overdue,
                    "rate":      prev_rate,
                    "penalty":   penalty,
                })

        cum_planned += planned_today
        cum_actual  += pay_by_date.get(curr_date, 0.0)

        prev_overdue = max(0.0, cum_planned - cum_actual)
        prev_rate    = get_cbr_rate(curr_date, cbr_rates)
        # Пени начисляются со дня, СЛЕДУЮЩЕГО за днём наступления срока платежа
        if planned_today > 0:
            prev_date = curr_date + timedelta(days=1)
        else:
            prev_date = curr_date

    return results


# ─── ПАРСЕРЫ EXCEL ────────────────────────────────────────────────────────────
_RU_MONTHS = {
    'января': 1, 'февраля': 2, 'марта': 3, 'апреля': 4,
    'мая': 5, 'июня': 6, 'июля': 7, 'августа': 8,
    'сентября': 9, 'октября': 10, 'ноября': 11, 'декабря': 12,
}

def _parse_date(val) -> Optional[date]:
    """Парсит дату: число, datetime, строку «31 мая 2021» или «31.05.2021»."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, (datetime,)):
        d = val.date()
        return d if d.year >= 2000 else None
    if hasattr(val, 'date'):  # pd.Timestamp
        d = val.date()
        return d if d.year >= 2000 else None
    s = str(val).strip()
    # «31 мая 2021»
    parts = s.split()
    if len(parts) == 3 and parts[1].lower() in _RU_MONTHS:
        try:
            d = date(int(parts[2]), _RU_MONTHS[parts[1].lower()], int(parts[0]))
            return d if d.year >= 2000 else None
        except ValueError:
            pass
    # «31.05.2021» или другие форматы
    try:
        d = pd.to_datetime(s, dayfirst=True).date()
        return d if d.year >= 2000 else None
    except Exception:
        return None


def _parse_amount(val) -> Optional[float]:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, (int, float)):
        return float(val) if val > 0 else None
    s = str(val).replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        v = float(s)
        return v if v > 0 else None
    except ValueError:
        return None


def parse_simple_excel(file) -> list[tuple[date, float]]:
    """Простой формат: колонка Дата | колонка Сумма."""
    df = pd.read_excel(file, header=0)
    result = []
    date_col = df.columns[0]
    amount_col = df.columns[1]
    for _, row in df.iterrows():
        d = _parse_date(row[date_col])
        if d is None:
            continue
        a = _parse_amount(row[amount_col])
        if a:
            result.append((d, a))
    return result


def parse_monthly_schedule_excel(file) -> list[tuple[date, float]]:
    """
    Ежемесячный график с заголовками: месяц | остаток | основной платеж | проценты | всего.
    Автоматически находит колонку с датой и колонку «основной платеж».
    """
    raw = pd.read_excel(file, header=None)

    date_col_idx: Optional[int] = None
    amount_col_idx: Optional[int] = None
    header_row_idx: Optional[int] = None

    skip_words = ["итого", "всего", "total", "полная"]

    # Найти строку-заголовок по ключевым словам
    for i, row in raw.iterrows():
        vals = [str(v).strip().lower() if pd.notna(v) else "" for v in row]
        row_text = " ".join(vals)
        if ("месяц" in row_text or "дата" in row_text) and "платеж" in row_text:
            header_row_idx = i
            for j, v in enumerate(vals):
                if ("месяц" in v or "дата" in v) and date_col_idx is None:
                    date_col_idx = j
                if "основной" in v and amount_col_idx is None:
                    amount_col_idx = j
            break

    # Если заголовок не найден — первая строка = заголовок, col0=дата, col1=сумма
    if date_col_idx is None:
        date_col_idx = 0
    if amount_col_idx is None:
        amount_col_idx = 1

    result: list[tuple[date, float]] = []
    for i, row in raw.iterrows():
        if header_row_idx is not None and i <= header_row_idx:
            continue

        date_val = row.iloc[date_col_idx] if date_col_idx < len(row) else None
        if date_val is None or (isinstance(date_val, float) and pd.isna(date_val)):
            continue

        # Пропускаем строки «Итого за год» и т.п.
        if any(kw in str(date_val).lower() for kw in skip_words):
            continue

        d = _parse_date(date_val)
        if d is None:
            continue

        amount_val = row.iloc[amount_col_idx] if amount_col_idx < len(row) else None
        a = _parse_amount(amount_val)
        if a:
            result.append((d, a))

    return result


def parse_pdf_schedule(file) -> list[tuple[date, float]]:
    """
    Извлекает график платежей из PDF с таблицей.
    Ищет колонки «месяц/дата» и «основной платеж/сумма».
    """
    import pdfplumber

    all_rows: list[list] = []
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if table:
                    all_rows.extend(table)

    if not all_rows:
        return []

    date_col_idx: Optional[int] = None
    amount_col_idx: Optional[int] = None
    header_row_idx: Optional[int] = None
    skip_words = ["итого", "всего", "total", "полная"]

    for i, row in enumerate(all_rows):
        if not row:
            continue
        vals = [str(v).strip().lower() if v else "" for v in row]
        row_text = " ".join(vals)
        if ("месяц" in row_text or "дата" in row_text) and ("платеж" in row_text or "сумма" in row_text):
            header_row_idx = i
            for j, v in enumerate(vals):
                if ("месяц" in v or "дата" in v) and date_col_idx is None:
                    date_col_idx = j
                if "основной" in v and amount_col_idx is None:
                    amount_col_idx = j
            if amount_col_idx is None:
                for j, v in enumerate(vals):
                    if ("платеж" in v or "сумма" in v) and j != date_col_idx and amount_col_idx is None:
                        amount_col_idx = j
            break

    if date_col_idx is None:
        date_col_idx = 0
    if amount_col_idx is None:
        amount_col_idx = 1

    result: list[tuple[date, float]] = []
    for i, row in enumerate(all_rows):
        if header_row_idx is not None and i <= header_row_idx:
            continue
        if not row or len(row) <= max(date_col_idx, amount_col_idx):
            continue
        date_val = row[date_col_idx]
        if date_val and any(kw in str(date_val).lower() for kw in skip_words):
            continue
        d = _parse_date(date_val)
        if d is None:
            continue
        a = _parse_amount(row[amount_col_idx])
        if a:
            result.append((d, a))

    return result


def parse_1c_excel(file) -> list[tuple[date, float]]:
    """
    Формат 1С «Анализ субконто».
    Ищет строки «Обороты за ДД.ММ.ГГ» в любом столбце,
    берёт значение из колонки Кредит (Обороты) или первое ненулевое число.
    """
    raw = pd.read_excel(file, header=None)

    # Найти колонку «Обороты Кредит»: в строке-заголовке ищем вторую «Кредит»
    kredit_oborot_col: Optional[int] = None
    for _, row in raw.iterrows():
        vals = [str(v).strip() if pd.notna(v) else "" for v in row]
        k_cols = [i for i, v in enumerate(vals) if "редит" in v or "РЕДИТ" in v]
        if len(k_cols) >= 2:
            kredit_oborot_col = k_cols[1]
            break

    payments: list[tuple[date, float]] = []
    date_re = re.compile(r"Обороты за\s+(\d{2}\.\d{2}\.\d{2,4})", re.IGNORECASE)

    for _, row in raw.iterrows():
        # Ищем «Обороты за» в ЛЮБОЙ ячейке строки (1С может вкладывать иерархически)
        ds: Optional[str] = None
        for cell_val in row:
            if pd.isna(cell_val):
                continue
            m = date_re.search(str(cell_val))
            if m:
                ds = m.group(1)
                break
        if ds is None:
            continue

        try:
            fmt = "%d.%m.%y" if len(ds) == 8 else "%d.%m.%Y"
            d = datetime.strptime(ds, fmt).date()
        except ValueError:
            continue

        amount: Optional[float] = None
        if kredit_oborot_col is not None and kredit_oborot_col < len(row):
            amount = _parse_amount(row.iloc[kredit_oborot_col])

        if amount is None:
            # Резервный вариант: берём первое ненулевое число в строке
            for v in row:
                a = _parse_amount(v)
                if a and a > 0:
                    amount = a
                    break

        if amount:
            payments.append((d, amount))

    return payments


# ─── ЭКСПОРТ В EXCEL (формат для суда) ───────────────────────────────────────
def export_excel(results: list[dict], info: dict, total: float) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Расчёт пеней"

    thin = Side(style="thin")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")

    # Заголовок
    ws.merge_cells("A1:G1")
    ws["A1"] = "РАСЧЁТ ПЕНИ ПО ДОГОВОРУ ЗАЙМА"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = center

    # Реквизиты договора
    r = 3
    for label, val in [
        ("Заёмщик:",         info["borrower"]),
        ("Договор №:",       f"{info['contract_num']} от {info['contract_date']}"),
        ("Дата расчёта:",    info["calc_date"]),
        ("Формула:",         "Сумма долга × 1/300 × Ключевая ставка ЦБ × Кол-во дней"),
    ]:
        ws.cell(r, 1, label).font = bold
        ws.merge_cells(f"B{r}:G{r}")
        ws.cell(r, 2, val)
        r += 1
    r += 1

    # Шапка таблицы
    headers = ["Период с", "Период по", "Дней", "Сумма долга, руб.", "Ставка ЦБ", "1/300 ставки", "Пени, руб."]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r, c, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = brd
    ws.row_dimensions[r].height = 30
    r += 1

    # Строки данных
    alt_fill = PatternFill("solid", fgColor="EBF0F8")
    for i, row in enumerate(results):
        fill = alt_fill if i % 2 == 0 else None
        data = [
            row["date_from"].strftime("%d.%m.%Y"),
            row["date_to"].strftime("%d.%m.%Y"),
            row["days"],
            round(row["debt"], 2),
            f"{row['rate']:.2f}%",
            f"{row['rate'] / 300:.4f}%",
            round(row["penalty"], 2),
        ]
        for c, val in enumerate(data, 1):
            cell = ws.cell(r, c, val)
            cell.border = brd
            if fill:
                cell.fill = fill
            cell.alignment = center if c in (1, 2, 3, 5, 6) else right
        r += 1

    # Итого
    ws.merge_cells(f"A{r}:F{r}")
    ws.cell(r, 1, "ИТОГО пеней:").font = bold
    ws.cell(r, 1).alignment = Alignment(horizontal="right")
    ws.cell(r, 7, round(total, 2)).font = Font(bold=True, size=11)
    ws.cell(r, 7).alignment = right
    for c in range(1, 8):
        ws.cell(r, c).border = brd
    r += 2

    ws.cell(r, 1, "* Ключевые ставки ЦБ РФ — cbr.ru. Проверяйте актуальность на дату расчёта.")
    ws.cell(r, 1).font = Font(italic=True, size=8, color="666666")
    ws.merge_cells(f"A{r}:G{r}")

    # Ширины колонок
    for c, w in enumerate([13, 13, 8, 22, 14, 15, 16], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── ШАБЛОН ГРАФИКА ──────────────────────────────────────────────────────────
def make_schedule_template() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "График платежей"
    ws["A1"], ws["B1"] = "Дата платежа", "Сумма платежа (руб.)"
    ws["A1"].font = ws["B1"].font = Font(bold=True)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 24
    # Пример
    for row, (d, a) in enumerate([
        ("31.03.2020", 15000.00),
        ("30.06.2020", 15000.00),
        ("30.09.2020", 15000.00),
        ("31.12.2020", 15000.00),
    ], start=2):
        ws.cell(row, 1, d)
        ws.cell(row, 2, a)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── UI ───────────────────────────────────────────────────────────────────────
st.title("⚖️ Калькулятор пеней по займам")
st.caption("АО ФАПК «Туймаада» | Пени = Сумма долга × 1/300 × Ставка ЦБ × Дней просрочки")

schedule_data: list[tuple[date, float]] = []
payments_data: list[tuple[date, float]] = []

tab1, tab2, tab3, tab4 = st.tabs(
    ["📋 Договор", "📅 График платежей", "💳 Оплаты", "🧮 Расчёт"]
)

# ── ТАБ 1: ДОГОВОР ────────────────────────────────────────────────────────────
with tab1:
    st.subheader("Данные договора")
    c1, c2 = st.columns(2)
    with c1:
        borrower    = st.text_input("ФИО заёмщика",    placeholder="Осипов Алексей Акимович")
        contract_num = st.text_input("Номер договора", placeholder="32-11-19")
    with c2:
        contract_date = st.date_input("Дата договора", value=date(2020, 3, 19),
                                      format="DD.MM.YYYY")
        calc_date = st.date_input("Дата расчёта пеней", value=date.today(),
                                  format="DD.MM.YYYY")

    st.info(
        "💡 Ставки ЦБ встроены с 2013 года. Если нужно обновить — "
        "отредактируйте таблицу на вкладке «Расчёт» (раздел «Ставки ЦБ»)."
    )

# ── ТАБ 2: ГРАФИК ─────────────────────────────────────────────────────────────
with tab2:
    st.subheader("График платежей по договору")

    sched_mode = st.radio(
        "Способ ввода:",
        ["📅 Квартальный (по годам)",
         "📂 Excel ежемесячный (месяц | основной платеж)",
         "📂 Excel простой (Дата | Сумма)",
         "📄 PDF (таблица с датами и суммами)"],
        horizontal=True,
        key="sched_mode",
    )

    if sched_mode == "📅 Квартальный (по годам)":
        q_input = st.radio(
            "Источник:",
            ["✏️ Ввести параметры вручную", "📂 Загрузить Excel (Дата | Сумма)"],
            horizontal=True,
            key="q_input",
        )

        if q_input == "✏️ Ввести параметры вручную":
            st.caption("Фиксированная сумма каждый квартал: 31 марта, 30 июня, 30 сентября, 31 декабря.")
            c1, c2, c3 = st.columns(3)
            with c1:
                q_year_start = st.number_input("Год начала", min_value=2000, max_value=2050,
                                               value=None, step=1, placeholder="например, 2020")
            with c2:
                q_year_end = st.number_input("Год окончания", min_value=2000, max_value=2050,
                                             value=None, step=1, placeholder="например, 2030")
            with c3:
                q_amount = st.number_input("Сумма за квартал (руб.)", min_value=0.0,
                                           value=None, step=0.01, format="%.2f",
                                           placeholder="например, 5721.03")

            if q_year_start and q_year_end and q_amount and q_amount > 0:
                quarter_ends = [(3, 31), (6, 30), (9, 30), (12, 31)]
                for yr in range(int(q_year_start), int(q_year_end) + 1):
                    for mon, day in quarter_ends:
                        schedule_data.append((date(yr, mon, day), float(q_amount)))
            else:
                st.info("Заполните все три поля для формирования графика.")

        else:
            st.caption("Формат файла: первая колонка — дата платежа, вторая — сумма. "
                       "Подходит для квартальных графиков с разными суммами.")
            c1, c2 = st.columns([3, 1])
            with c1:
                q_file = st.file_uploader("Excel-файл квартального графика",
                                          type=["xlsx", "xls"], key="sched_q_excel")
            with c2:
                st.download_button(
                    "📥 Шаблон",
                    data=make_schedule_template(),
                    file_name="шаблон_квартальный.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            if q_file:
                try:
                    schedule_data = parse_simple_excel(q_file)
                except Exception as e:
                    st.error(f"Ошибка чтения: {e}")

        if schedule_data:
            df_s = pd.DataFrame(schedule_data, columns=["Дата", "Сумма (руб.)"])
            df_s["Дата"] = df_s["Дата"].apply(lambda d: d.strftime("%d.%m.%Y"))
            df_s["Сумма (руб.)"] = df_s["Сумма (руб.)"].apply(lambda x: f"{x:,.2f}")
            st.success(f"✅ {len(schedule_data)} платежей | "
                       f"Итого: {sum(a for _, a in schedule_data):,.2f} ₽")
            st.dataframe(df_s, use_container_width=True, hide_index=True)

    elif sched_mode == "📂 Excel ежемесячный (месяц | основной платеж)":
        st.caption("Формат ежемесячного графика: колонки «месяц», «остаток», «основной платеж», «проценты», «всего».")
        schedule_file = st.file_uploader(
            "Excel-файл графика", type=["xlsx", "xls"], key="sched_monthly"
        )
        if schedule_file:
            try:
                schedule_data = parse_monthly_schedule_excel(schedule_file)
                if schedule_data:
                    df_s = pd.DataFrame(schedule_data, columns=["Дата", "Сумма (руб.)"])
                    df_s["Дата"] = df_s["Дата"].apply(lambda d: d.strftime("%d.%m.%Y"))
                    df_s["Сумма (руб.)"] = df_s["Сумма (руб.)"].apply(lambda x: f"{x:,.2f}")
                    st.success(f"✅ {len(schedule_data)} платежей | "
                               f"Итого: {sum(a for _, a in schedule_data):,.2f} ₽")
                    st.dataframe(df_s, use_container_width=True, hide_index=True)
                else:
                    # Показываем первые строки файла для диагностики
                    raw_diag = pd.read_excel(schedule_file, header=None)
                    st.warning("Платежи не найдены. Первые строки файла:")
                    st.dataframe(raw_diag.head(10), use_container_width=True)
            except Exception as e:
                st.error(f"Ошибка чтения: {e}")
        else:
            st.info("Загрузите Excel-файл ежемесячного графика (например, выгруженный из расчёта займа).")

    elif sched_mode == "📂 Excel простой (Дата | Сумма)":
        c1, c2 = st.columns([3, 1])
        with c1:
            schedule_file = st.file_uploader(
                "Excel-файл графика", type=["xlsx", "xls"], key="sched"
            )
        with c2:
            st.download_button(
                "📥 Шаблон",
                data=make_schedule_template(),
                file_name="шаблон_график.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        if schedule_file:
            try:
                schedule_data = parse_simple_excel(schedule_file)
                if schedule_data:
                    df_s = pd.DataFrame(schedule_data, columns=["Дата", "Сумма (руб.)"])
                    df_s["Дата"] = df_s["Дата"].apply(lambda d: d.strftime("%d.%m.%Y"))
                    df_s["Сумма (руб.)"] = df_s["Сумма (руб.)"].apply(lambda x: f"{x:,.2f}")
                    st.success(f"✅ {len(schedule_data)} платежей | "
                               f"Итого: {sum(a for _, a in schedule_data):,.2f} ₽")
                    st.dataframe(df_s, use_container_width=True, hide_index=True)
                else:
                    st.warning("Платежи не найдены. Проверьте формат файла.")
            except Exception as e:
                st.error(f"Ошибка чтения: {e}")
        else:
            st.info("Формат файла: первая колонка — дата, вторая — сумма. Скачайте шаблон →")

    else:
        st.caption("Подходит для PDF-графиков из банков и расчётных систем. "
                   "Программа автоматически найдёт таблицу с датами и суммами.")
        schedule_file = st.file_uploader(
            "PDF-файл графика", type=["pdf"], key="sched_pdf"
        )
        if schedule_file:
            try:
                schedule_data = parse_pdf_schedule(schedule_file)
                if schedule_data:
                    df_s = pd.DataFrame(schedule_data, columns=["Дата", "Сумма (руб.)"])
                    df_s["Дата"] = df_s["Дата"].apply(lambda d: d.strftime("%d.%m.%Y"))
                    df_s["Сумма (руб.)"] = df_s["Сумма (руб.)"].apply(lambda x: f"{x:,.2f}")
                    st.success(f"✅ {len(schedule_data)} платежей | "
                               f"Итого: {sum(a for _, a in schedule_data):,.2f} ₽")
                    st.dataframe(df_s, use_container_width=True, hide_index=True)
                else:
                    raw_diag: list[list] = []
                    import pdfplumber
                    with pdfplumber.open(schedule_file) as pdf:
                        for page in pdf.pages:
                            for tbl in page.extract_tables():
                                raw_diag.extend(tbl)
                    st.warning("Таблица не распознана. Первые строки из PDF:")
                    st.dataframe(pd.DataFrame(raw_diag[:10]), use_container_width=True)
            except Exception as e:
                st.error(f"Ошибка чтения PDF: {e}")
        else:
            st.info("Загрузите PDF-файл графика платежей. Программа сама найдёт столбцы с датами и суммами.")

# ── ТАБ 3: ОПЛАТЫ ─────────────────────────────────────────────────────────────
with tab3:
    st.subheader("Фактические оплаты")

    fmt = st.radio(
        "Формат файла:",
        ["Анализ субконто (1С)", "Простой формат (Дата | Сумма)"],
        horizontal=True,
    )
    payments_file = st.file_uploader(
        "Excel-файл оплат", type=["xlsx", "xls"], key="pays"
    )

    if payments_file:
        try:
            if "1С" in fmt:
                payments_data = parse_1c_excel(payments_file)
            else:
                payments_data = parse_simple_excel(payments_file)

            if not payments_data and "1С" in fmt:
                raw_diag = pd.read_excel(payments_file, header=None)
                st.warning("Оплаты не найдены. Первые строки файла (для диагностики):")
                st.dataframe(raw_diag.head(12), use_container_width=True)
                st.info("Если столбцы выглядят правильно — попробуйте «Простой формат (Дата | Сумма)».")
            elif payments_data:
                df_p = pd.DataFrame(payments_data, columns=["Дата", "Сумма (руб.)"])
                df_p["Дата"] = df_p["Дата"].apply(lambda d: d.strftime("%d.%m.%Y"))
                total_paid = sum(a for _, a in payments_data)
                df_p["Сумма (руб.)"] = df_p["Сумма (руб.)"].apply(lambda x: f"{x:,.2f}")
                st.success(f"✅ {len(payments_data)} оплат | Итого: {total_paid:,.2f} ₽")
                st.dataframe(df_p, use_container_width=True, hide_index=True)
            else:
                st.warning("Оплаты не найдены. Если используете 1С — выберите соответствующий формат.")
        except Exception as e:
            st.error(f"Ошибка чтения: {e}")
    else:
        st.info(
            "**1С:** Файл → Сохранить как → Excel из «Анализ субконто».\n\n"
            "**Простой формат:** Дата | Сумма оплаты."
        )

# ── ТАБ 4: РАСЧЁТ ─────────────────────────────────────────────────────────────
with tab4:
    st.subheader("Расчёт пеней")

    with st.expander("📊 Таблица ключевых ставок ЦБ (редактируемая)"):
        df_rates = pd.DataFrame(CBR_RATES, columns=["Дата вступления", "Ставка (%)"])
        df_rates["Дата вступления"] = df_rates["Дата вступления"].apply(
            lambda d: d.strftime("%d.%m.%Y")
        )
        edited = st.data_editor(df_rates, use_container_width=True, hide_index=True,
                                num_rows="dynamic")
        # Парсим обратно
        custom_rates: list[tuple[date, float]] = []
        for _, row in edited.iterrows():
            try:
                d = datetime.strptime(str(row["Дата вступления"]), "%d.%m.%Y").date()
                r = float(row["Ставка (%)"])
                custom_rates.append((d, r))
            except Exception:
                pass
        cbr_to_use = sorted(custom_rates, key=lambda x: x[0]) if custom_rates else CBR_RATES

    if st.button("🧮 Рассчитать пени", type="primary", use_container_width=True):
        if not schedule_data:
            st.error("⚠️ Загрузите график платежей на вкладке «График платежей»")
        else:
            results = calculate_penalties(schedule_data, payments_data, cbr_to_use, calc_date)

            if not results:
                st.success("✅ Просрочек не обнаружено!")
            else:
                total_penalty  = sum(r["penalty"] for r in results)
                total_planned  = sum(a for d, a in schedule_data if d <= calc_date)
                total_paid_sum = sum(a for d, a in payments_data if d <= calc_date)
                overdue_debt   = max(0.0, total_planned - total_paid_sum)

                m1, m2, m3 = st.columns(3)
                m1.metric("По графику (накоплено)", f"{total_planned:,.2f} ₽")
                m2.metric("Фактически оплачено",     f"{total_paid_sum:,.2f} ₽")
                m3.metric("Сумма пеней",              f"{total_penalty:,.2f} ₽",
                          delta=f"Остаток долга: {overdue_debt:,.2f} ₽",
                          delta_color="inverse")

                # Таблица расчёта
                df_res = pd.DataFrame([{
                    "Период с":           r["date_from"].strftime("%d.%m.%Y"),
                    "Период по":          r["date_to"].strftime("%d.%m.%Y"),
                    "Дней":               r["days"],
                    "Сумма долга, руб.":  f"{r['debt']:,.2f}",
                    "Ставка ЦБ":          f"{r['rate']:.2f}%",
                    "1/300 ставки":       f"{r['rate']/300:.4f}%",
                    "Пени, руб.":         f"{r['penalty']:,.2f}",
                } for r in results])
                st.dataframe(df_res, use_container_width=True, hide_index=True)
                st.markdown(f"**Итого пеней: {total_penalty:,.2f} ₽**")

                # Скачать Excel
                info = {
                    "borrower":      borrower or "—",
                    "contract_num":  contract_num or "—",
                    "contract_date": contract_date.strftime("%d.%m.%Y"),
                    "calc_date":     calc_date.strftime("%d.%m.%Y"),
                }
                st.download_button(
                    "📥 Скачать расчёт для суда (Excel)",
                    data=export_excel(results, info, total_penalty),
                    file_name=f"пени_{contract_num or 'договор'}_{calc_date}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                )
